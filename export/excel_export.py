"""
Módulo de exportação para Excel
"""
from typing import List, Dict, Any, Optional
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference
import numpy as np
from pathlib import Path

from core.game import Game
from core.statistics import StatisticsAnalyzer
from core.population import Population


class ExcelExporter:
    """
    Exportador para Excel com formatação profissional
    
    Gera:
    - Planilha de jogos
    - Estatísticas
    - Frequências
    - Análises
    - Gráficos
    """
    
    def __init__(self):
        self.wb = None
        self.stats_analyzer = StatisticsAnalyzer()
        
    def export(self, games: List[Game], filename: str, config: Dict[str, Any]):
        """
        Exporta jogos para arquivo Excel
        
        Args:
            games: Lista de jogos
            filename: Nome do arquivo
            config: Configurações do sistema
        """
        self.wb = Workbook()
        
        # Remove planilha padrão
        self.wb.remove(self.wb.active)
        
        # Cria planilhas
        self._create_games_sheet(games)
        self._create_statistics_sheet(games)
        self._create_frequency_sheet(games)
        self._create_analysis_sheet(games)
        
        # Salva
        self.wb.save(filename)
    
    def _create_games_sheet(self, games: List[Game]):
        """Cria planilha com os jogos"""
        ws = self.wb.create_sheet("Jogos", 0)
        
        # Cabeçalho
        headers = ["Jogo"] + [f"N{i+1}" for i in range(15)] + ["Soma", "Pares", "Ímpares", "Score"]
        ws.append(headers)
        
        # Estilo do cabeçalho
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Dados
        for i, game in enumerate(games, 1):
            row = [f"Jogo {i}"]
            row.extend(game.numbers)
            row.extend([game.soma, game.pares, game.impares, game.score])
            ws.append(row)
        
        # Ajusta largura das colunas
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws.column_dimensions[column].width = adjusted_width
        
        # Formata números
        for row in range(2, len(games) + 2):
            for col in range(2, 17):  # Colunas dos números
                cell = ws.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal="center")
            
            # Score com 4 casas decimais
            score_cell = ws.cell(row=row, column=len(headers))
            score_cell.number_format = "0.0000"
    
    def _create_statistics_sheet(self, games: List[Game]):
        """Cria planilha com estatísticas"""
        ws = self.wb.create_sheet("Estatísticas")
        
        # Análise estatística
        stats = self.stats_analyzer.analyze_population(Population(games))
        
        # Estatísticas básicas
        row = 1
        ws.cell(row=row, column=1, value="ESTATÍSTICAS BÁSICAS")
        ws.cell(row=row, column=1).font = Font(bold=True, size=14)
        row += 2
        
        stats_data = [
            ("Total de Jogos", len(games)),
            ("Score Médio", f"{sum(g.score for g in games) / len(games):.4f}" if games else "0"),
            ("Melhor Score", f"{max(g.score for g in games):.4f}" if games else "0"),
            ("Pior Score", f"{min(g.score for g in games):.4f}" if games else "0"),
            ("Desvio Padrão Score", f"{np.std([g.score for g in games]):.4f}" if games else "0"),
            ("Frequência STD", f"{stats.frequency_std:.4f}"),
            ("Entropia", f"{stats.frequency_entropy:.4f}"),
        ]
        
        for label, value in stats_data:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=value)
            row += 1
        
        # Distribuição de Pares
        row += 2
        ws.cell(row=row, column=1, value="DISTRIBUIÇÃO DE PARES")
        ws.cell(row=row, column=1).font = Font(bold=True, size=14)
        row += 2
        
        ws.cell(row=row, column=1, value="Quantidade de Pares")
        ws.cell(row=row, column=2, value="Frequência")
        
        for pairs, freq in sorted(stats.pair_distribution.items()):
            row += 1
            ws.cell(row=row, column=1, value=pairs)
            ws.cell(row=row, column=2, value=freq)
        
        # Distribuição de Somas
        row += 2
        ws.cell(row=row, column=1, value="DISTRIBUIÇÃO DE SOMAS")
        ws.cell(row=row, column=1).font = Font(bold=True, size=14)
        row += 2
        
        ws.cell(row=row, column=1, value="Faixa de Soma")
        ws.cell(row=row, column=2, value="Frequência")
        
        for sum_range, freq in sorted(stats.sum_distribution.items()):
            row += 1
            ws.cell(row=row, column=1, value=f"{sum_range}-{sum_range + 9}")
            ws.cell(row=row, column=2, value=freq)
        
        # Formatação
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 30)
    
    def _create_frequency_sheet(self, games: List[Game]):
        """Cria planilha com frequências"""
        ws = self.wb.create_sheet("Frequências")
        
        # Cabeçalho
        ws.cell(row=1, column=1, value="Dezena")
        ws.cell(row=1, column=2, value="Frequência")
        ws.cell(row=1, column=3, value="Porcentagem")
        
        # Estilo cabeçalho
        for col in range(1, 4):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Dados
        freq = [0] * 25
        for game in games:
            for num in game.numbers:
                freq[num - 1] += 1
        
        total = len(games) * 15
        for i, count in enumerate(freq, 1):
            row = i + 1
            ws.cell(row=row, column=1, value=i)
            ws.cell(row=row, column=2, value=count)
            ws.cell(row=row, column=3, value=f"{count / total * 100:.2f}%")
        
        # Formatação
        for row in range(2, 27):
            ws.cell(row=row, column=2).alignment = Alignment(horizontal="center")
            ws.cell(row=row, column=3).alignment = Alignment(horizontal="center")
        
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        
        # Gráfico de barras
        chart = BarChart()
        chart.title = "Frequência das Dezenas"
        chart.x_axis.title = "Dezena"
        chart.y_axis.title = "Frequência"
        
        data = Reference(ws, min_col=2, min_row=1, max_row=26, max_col=2)
        categories = Reference(ws, min_col=1, min_row=2, max_row=26)
        
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.width = 15
        chart.height = 10
        
        ws.add_chart(chart, "E2")
    
    def _create_analysis_sheet(self, games: List[Game]):
        """Cria planilha com análises detalhadas"""
        ws = self.wb.create_sheet("Análises")
        
        row = 1
        
        # Análise de Linhas
        ws.cell(row=row, column=1, value="DISTRIBUIÇÃO POR LINHAS")
        ws.cell(row=row, column=1).font = Font(bold=True, size=14)
        row += 2
        
        ws.cell(row=row, column=1, value="Linha")
        ws.cell(row=row, column=2, value="Média")
        ws.cell(row=row, column=3, value="Mínimo")
        ws.cell(row=row, column=4, value="Máximo")
        ws.cell(row=row, column=5, value="Desvio Padrão")
        
        row += 1
        for line in range(5):
            values = [game.linhas[line] for game in games]
            ws.cell(row=row, column=1, value=f"Linha {line + 1}")
            ws.cell(row=row, column=2, value=f"{np.mean(values):.2f}")
            ws.cell(row=row, column=3, value=min(values))
            ws.cell(row=row, column=4, value=max(values))
            ws.cell(row=row, column=5, value=f"{np.std(values):.2f}")
            row += 1
        
        # Análise de Colunas
        row += 2
        ws.cell(row=row, column=1, value="DISTRIBUIÇÃO POR COLUNAS")
        ws.cell(row=row, column=1).font = Font(bold=True, size=14)
        row += 2
        
        ws.cell(row=row, column=1, value="Coluna")
        ws.cell(row=row, column=2, value="Média")
        ws.cell(row=row, column=3, value="Mínimo")
        ws.cell(row=row, column=4, value="Máximo")
        ws.cell(row=row, column=5, value="Desvio Padrão")
        
        row += 1
        for col in range(5):
            values = [game.colunas[col] for game in games]
            ws.cell(row=row, column=1, value=f"Coluna {col + 1}")
            ws.cell(row=row, column=2, value=f"{np.mean(values):.2f}")
            ws.cell(row=row, column=3, value=min(values))
            ws.cell(row=row, column=4, value=max(values))
            ws.cell(row=row, column=5, value=f"{np.std(values):.2f}")
            row += 1
        
        # Análise de Consecutivos
        row += 2
        ws.cell(row=row, column=1, value="ANÁLISE DE CONSECUTIVOS")
        ws.cell(row=row, column=1).font = Font(bold=True, size=14)
        row += 2
        
        ws.cell(row=row, column=1, value="Consecutivos")
        ws.cell(row=row, column=2, value="Frequência")
        ws.cell(row=row, column=3, value="Porcentagem")
        
        consec_stats = self.stats_analyzer.analyze_population(Population(games)).consecutive_analysis
        total = len(games)
        
        for consec, freq in sorted(consec_stats.items()):
            row += 1
            ws.cell(row=row, column=1, value=consec)
            ws.cell(row=row, column=2, value=freq)
            ws.cell(row=row, column=3, value=f"{freq / total * 100:.1f}%")
        
        # Formatação
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 20)
